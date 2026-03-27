"""
Generate a list of Active, non-life members who have NOT paid their
2026 Membership (eventid = 149).

Run inside the backend container:
  docker exec bcs_backend python /app/generate_2026_unpaid_report.py

Output: /app/2026_unpaid_members.xlsx
"""

import os
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

DB_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://bcs_user:bcs_password@db:5432/bcs_registration",
)
OUTPUT = "/app/2026_unpaid_members.xlsx"

QUERY = """
SELECT
    m.personid,
    m.firstname,
    m.lastname,
    COALESCE(m.email, '')      AS email,
    COALESCE(m.cellphone, '')  AS cellphone,
    COALESCE(m.homephone, '')  AS homephone,
    COALESCE(m.address1, '')   AS address1,
    COALESCE(m.address2, '')   AS address2,
    COALESCE(m.city, '')       AS city,
    COALESCE(m.state, '')      AS state,
    COALESCE(m.zip, '')        AS zip
FROM bcs_members m
WHERE m.status = 'Active'
  AND (m.lifemember = FALSE OR m.lifemember IS NULL)
  AND m.personid NOT IN (
      SELECT DISTINCT c.personid
      FROM bcs_contributions c
      WHERE c.eventid = 149
  )
ORDER BY m.lastname, m.firstname;
"""

# ── colours ───────────────────────────────────────────────────────────────────
HEADER_BG   = "1F3A6E"   # BCS dark blue
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "EEF2F8"   # light blue-grey
TOTAL_BG    = "D9E1F2"

THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def col_border(cell):
    cell.border = BORDER

def make_header(ws, headers, col_widths):
    ws.append(headers)
    for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

def style_data_row(ws, row_num, num_cols, alternate):
    fill = PatternFill("solid", fgColor=ALT_ROW_BG) if alternate else None
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border    = BORDER
        if fill:
            cell.fill = fill

def main():
    print("Connecting to PostgreSQL…")
    conn = psycopg2.connect(DB_URL)
    cur  = conn.cursor()

    cur.execute(QUERY)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    print(f"Found {len(rows)} members who have NOT paid their 2026 Membership.")

    wb = Workbook()
    ws = wb.active
    ws.title = "2026 Unpaid Members"

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value     = "Bengali Cultural Society — 2026 Membership Dues Unpaid"
    title_cell.font      = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value     = f"Active Members (excluding Life Members) with no 2026 Membership contribution  ·  Total: {len(rows)}"
    sub_cell.font      = Font(name="Arial", italic=True, size=10, color="555555")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.append([])   # blank spacer row

    # ── Column headers (row 4) ────────────────────────────────────────────────
    headers    = ["#", "First Name", "Last Name", "Email", "Cell Phone",
                  "Home Phone", "Address 1", "Address 2", "City", "State", "Zip"]
    col_widths = [5, 14, 16, 30, 14, 14, 24, 18, 16, 7, 10]

    ws.row_dimensions[4].height = 28
    ws.append(headers)   # row 4

    for col_idx, (cell, width) in enumerate(zip(ws[4], col_widths), start=1):
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows (start at row 5) ────────────────────────────────────────────
    for i, (pid, fname, lname, email, cell_ph, home_ph,
            addr1, addr2, city, state, zip_) in enumerate(rows, start=1):
        row_num = i + 4   # offset: 3 header rows + 1 blank
        ws.append([i, fname, lname, email, cell_ph, home_ph,
                   addr1, addr2, city, state, zip_])
        alternate = (i % 2 == 0)
        fill = PatternFill("solid", fgColor=ALT_ROW_BG) if alternate else None
        for col in range(1, 12):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col in (1, 10, 11) else "left")
            cell.border    = BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_num].height = 16

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1).value = "Total"
    ws.cell(row=total_row, column=2).value = f"=COUNTA(B5:B{total_row-1})"
    for col in range(1, 12):
        cell = ws.cell(row=total_row, column=col)
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.fill      = PatternFill("solid", fgColor=TOTAL_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[total_row].height = 18

    # ── Freeze panes below header row ─────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A4:K{total_row - 1}"

    wb.save(OUTPUT)
    print(f"✅ Report saved to {OUTPUT}")
    print(f"   On your Mac: /Users/sdhar/development/bcs-registration/backend/2026_unpaid_members.xlsx")

if __name__ == "__main__":
    main()
